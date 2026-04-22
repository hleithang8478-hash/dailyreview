import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from config import FIELD_MAPPING
import re
import logging

# 获取logger
logger = logging.getLogger(__name__)

def translate_columns(df):
    """将DataFrame列名翻译为中文"""
    if df.empty:
        return df
    
    translated_columns = {}
    for col in df.columns:
        if col in FIELD_MAPPING:
            translated_columns[col] = FIELD_MAPPING[col]
        else:
            # 处理带后缀的字段
            for key, value in FIELD_MAPPING.items():
                if col.endswith(key):
                    suffix = col[len(key):]
                    if suffix == '_当前值':
                        translated_columns[col] = value + '-当前值'
                    elif suffix == '_分位数':
                        translated_columns[col] = value + '-分位数'
                    else:
                        translated_columns[col] = value + suffix
                    break
            else:
                # 如果没有找到匹配，保持原列名
                translated_columns[col] = col
    
    df_translated = df.copy()
    df_translated.columns = [translated_columns.get(col, col) for col in df.columns]
    return df_translated

def get_field_category(field_name):
    """获取字段分类和排序权重"""
    # 基础信息
    if any(keyword in field_name for keyword in ['股票代码', 'SecuCode', 'CompanyCode', 'InnerCode', '股票简称']):
        return '基础信息', 1
    
    # 行业信息
    if any(keyword in field_name for keyword in ['一级行业', '二级行业', '三级行业']):
        return '行业信息', 2
    
    # 日期信息
    if any(keyword in field_name for keyword in ['报告期', '行情日期']):
        return '日期信息', 3
    
    # 盈利能力
    if any(keyword in field_name for keyword in ['ROETTM', 'ROICTTM', 'NetProfitRatioTTM', 'NPToTORTTM', 'OperatingRevenueCashCover', 'NetProfitCashCover', 'MainProfitProportion']):
        return '盈利能力', 4
    
    # 费用控制
    if any(keyword in field_name for keyword in ['OperatingExpenseRateTTM', 'AdminiExpenseRateTTM', 'FinancialExpenseRateTTM']):
        return '费用控制', 5
    
    # 偿债能力
    if any(keyword in field_name for keyword in ['SuperQuickRatio', 'NOCFToCurrentLiability', 'NetAssetLiabilityRatio', 'InteBearDebtToTL']):
        return '偿债能力', 6
    
    # 成长能力
    if any(keyword in field_name for keyword in ['OperatingRevenueGrowRate', 'TORGrowRate', 'NetOperateCashFlowYOY', 'OperCashPSGrowRate']):
        return '成长能力', 7
    
    # 分红指标
    if any(keyword in field_name for keyword in ['DividendPS', 'DividendPaidRatio', 'DividendTTM']):
        return '分红指标', 8
    
    # 规模指标
    if any(keyword in field_name for keyword in ['TotalAssets', 'TotalEquity', 'OperatingRevenue', 'NetProfit']):
        return '规模指标', 8
    
    # 行情数据
    if any(keyword in field_name for keyword in ['收盘价', '成交量']):
        return '行情数据', 9
    
    # 趋势分析字段
    if any(keyword in field_name for keyword in ['5年CAGR', '趋势斜率', '趋势标签', '起始值', '当前值', '变化幅度']):
        return '趋势分析', 10
    
    # 历史数据字段
    if any(keyword in field_name for keyword in ['近5年数据', '年份']):
        return '历史数据', 11
    
    return '其他', 99

def sort_columns_by_category(df):
    """按分类对列进行排序"""
    if df.empty:
        return df
    
    # 获取每列的分类和权重
    column_categories = {}
    for col in df.columns:
        category, weight = get_field_category(col)
        column_categories[col] = (category, weight, col)  # 添加列名作为第三排序键
    
    # 按分类权重和列名排序
    sorted_columns = sorted(df.columns, key=lambda x: column_categories[x])
    
    return df[sorted_columns]

def format_excel_worksheet(ws, df, sheet_name):
    """格式化Excel工作表"""
    if df.empty:
        return
    
    # 写入数据
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # 获取数据范围
    max_row = ws.max_row
    max_col = ws.max_column
    
    # 设置列宽
    for col in range(1, max_col + 1):
        column_letter = ws.cell(row=1, column=col).column_letter
        max_length = 0
        for row in range(1, max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        # 设置列宽，最小10，最大50
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 格式化表头
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 格式化数据行
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            
            # 交替行颜色
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # 如果是综合汇总sheet，添加超链接
    if sheet_name == "综合汇总":
        add_hyperlinks_to_summary(ws, df)
    
    # 冻结首行
    ws.freeze_panes = "A2"

def add_hyperlinks_to_summary(ws, df):
    """为综合汇总sheet添加超链接"""
    try:
        from openpyxl.styles import Font
        
        # 查找股票代码列的索引
        stock_code_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == '股票代码':
                stock_code_col = col
                break
        
        if stock_code_col is None:
            return
        
        # 为每个股票代码添加超链接
        for row in range(2, ws.max_row + 1):
            stock_code = ws.cell(row=row, column=stock_code_col).value
            if stock_code:
                # 创建超链接到汇总sheet
                cell = ws.cell(row=row, column=stock_code_col)
                
                # 使用当前行号作为目标行号（假设汇总sheet中的行号相同）
                target_row = row
                cell.hyperlink = f"#汇总!A{target_row}"
                cell.font = Font(color="0000FF", underline="single")
                cell.value = stock_code
                
    except Exception as e:
        logger.info(f"添加超链接时出错: {e}")

def format_numeric_data(df):
    """格式化数值数据"""
    if df.empty:
        return df
    
    df_formatted = df.copy()
    
    # 需要保持字符串格式的字段（避免丢失前导零）
    string_fields = ['股票代码', 'SecuCode', 'CompanyCode', 'InnerCode', '行情日期', '报告期', 
                    '一级行业', '二级行业', '三级行业', '股票简称', '趋势标签', '指标分类', '指标名称']
    
    # 需要特殊格式化的字段
    for col in df_formatted.columns:
        if col in string_fields:
            # 字符串字段，将空值显示为"-"
            df_formatted[col] = df_formatted[col].fillna("-")
            continue
        
        # 数值字段处理
        if df_formatted[col].dtype in ['float64', 'int64'] or df_formatted[col].dtype == 'object':
            # 尝试转换为数值
            numeric_series = pd.to_numeric(df_formatted[col], errors='coerce')
            
            # 如果转换成功，应用格式化
            if not numeric_series.isna().all():
                # 分位数字段格式化为百分比
                if '分位数' in col or '%' in col:
                    df_formatted[col] = numeric_series.apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "-")
                # CAGR和变化幅度字段格式化为百分比
                elif any(keyword in col for keyword in ['CAGR', '变化幅度', '增长率']):
                    df_formatted[col] = numeric_series.apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "-")
                # 其他数值字段
                else:
                    df_formatted[col] = numeric_series.apply(lambda x: f"{x:.4f}" if pd.notna(x) else "-")
            else:
                # 转换失败，保持原值，空值显示为"-"
                df_formatted[col] = df_formatted[col].fillna("-")
    
    return df_formatted

def export_results(df, output_file, sheet_name="数据"):
    """
    导出结果到Excel文件 - 修改版本
    Args:
        df: 要导出的DataFrame
        output_file: 输出文件路径
        sheet_name: 工作表名称
    """
    if df.empty:
        logger.info(f"警告: {sheet_name} 数据为空，跳过导出")
        return
    
    try:
        # 检查并隐藏全空的列
        columns_to_remove = []
        for col in df.columns:
            # 检查列是否全为空
            if df[col].dtype in ['object', 'string']:
                # 对于字符串类型，检查是否全为空字符串或NaN
                if df[col].isna().all() or (df[col].astype(str) == '').all():
                    columns_to_remove.append(col)
            else:
                # 对于数值类型，检查是否全为0或NaN
                if df[col].isna().all() or (df[col] == 0).all():
                    columns_to_remove.append(col)
        
        # 移除全空的列
        if columns_to_remove:
            df = df.drop(columns=columns_to_remove)
        
        # 翻译列名
        df_translated = translate_columns(df)
        
        # 格式化数值数据
        df_formatted = format_numeric_data(df_translated)
        
        # 按分类排序列
        df_sorted = sort_columns_by_category(df_formatted)
        
        # 创建或加载工作簿
        try:
            from openpyxl import load_workbook
            wb = load_workbook(output_file)
        except FileNotFoundError:
            wb = Workbook()
            # 删除默认的Sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # 检查工作表是否已存在，如果存在则删除
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        
        # 创建工作表
        ws = wb.create_sheet(sheet_name)
        
        # 格式化工作表
        format_excel_worksheet(ws, df_sorted, sheet_name)
        
        # 保存文件
        wb.save(output_file)
        logger.info(f"成功导出 {sheet_name} 数据到 {output_file}")
        logger.info(f"  数据行数: {len(df_sorted)}")
        logger.info(f"  数据列数: {len(df_sorted.columns)}")
        
    except Exception as e:
        logger.info(f"导出 {sheet_name} 数据时出错: {e}")
        import traceback
        traceback.print_exc()

# 保持向后兼容的函数
def export_results_old(crash_stocks, fundamental_data, output_file):
    """向后兼容的导出函数"""
    try:
        # 创建或加载工作簿
        try:
            from openpyxl import load_workbook
            wb = load_workbook(output_file)
        except FileNotFoundError:
            wb = Workbook()
            # 删除默认的Sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # 导出大跌股票数据
        if not crash_stocks.empty:
            export_results(crash_stocks, output_file, "大跌股票筛选结果")
        
        # 导出基本面数据
        if isinstance(fundamental_data, tuple) and len(fundamental_data) == 3:
            current_df, trends_df, history_df = fundamental_data
            
            if not current_df.empty:
                export_results(current_df, output_file, "基本面分析-当前值")
            
            if not trends_df.empty:
                export_results(trends_df, output_file, "基本面分析-趋势分析")
            
            if not history_df.empty:
                export_results(history_df, output_file, "基本面分析-历史数据")
        elif not fundamental_data.empty:
            export_results(fundamental_data, output_file, "基本面分析")
        
        logger.info(f"所有数据已导出到: {output_file}")
        
    except Exception as e:
        logger.info(f"导出结果时出错: {e}")
        import traceback
        traceback.print_exc() 